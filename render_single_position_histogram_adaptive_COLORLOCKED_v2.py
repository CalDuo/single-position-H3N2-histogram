"""
Locked single-position histogram renderer (adaptive neutralization height) with:
- Canonical AA color map frozen from all_positions_part05.svg (historical reference)
- Antigenic cluster strip aligned to bins
- Neutralization matrix (shaded fills) + mutation letters
- Year axis placed directly beneath the neutralization block (no stretching)
- Histogram y-ticks fixed to 0.25 increments

Outputs: PNG or SVG (transparent background)
"""

from __future__ import annotations

import json
from pathlib import Path
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
import openpyxl

plt.rcParams["font.family"] = "DejaVu Sans"

# ---- CANONICAL AA COLOR MAP (DO NOT MODIFY) ----
AA_COLOR_MAP = {
    "A": "#e6194b",
    "C": "#3cb44b",
    "D": "#ffe119",
    "E": "#4363d8",
    "F": "#f58231",
    "G": "#911eb4",
    "H": "#46f0f0",
    "I": "#f032e6",
    "K": "#bcf60c",
    "L": "#fabebe",
    "M": "#808080",
    "N": "#008080",
    "P": "#e6beff",
    "Q": "#9a6324",
    "R": "#fffac8",
    "S": "#800000",
    "T": "#aaffc3",
    "V": "#808000",
    "W": "#ffd8b1",
    "Y": "#000075",
    "Other": "#BFBFBF"
}

# ---- LOCKED CLUSTER DEFINITIONS ----
CLUSTERS = [
    ("HK68",1968,1972),("EN72",1972,1975),("VI75",1975,1977),
    ("TX77",1977,1979),("BK79",1979,1987),("SI87",1987,1989),
    ("BE89",1989,1992),("BE92",1992,1995),("WU95",1995,1997),
    ("SY97",1997,2002),("FU02",2002,2004),("CA04",2004,2005),
    ("WI05",2005,2009),("PE09",2009,2012),("TX12",2012,2014),
    ("HK14",2014,2015)
]
CLUSTER_NAMES = [c[0] for c in CLUSTERS]
CLUSTER_COLORS = {n: plt.get_cmap("tab20")(i) for i,(n,_,_) in enumerate(CLUSTERS)}

# ---- LOCKED YEAR RANGE ----
YEAR_MIN, YEAR_MAX = 1965, 2020

# ---- LOCKED LAYOUT (inches) ----
HIST_H_IN = 2.0
CLUSTER_H_IN = 0.30
YEAR_H_IN = 0.25
GAP_IN = 0.10
TOP_PAD_IN = 0.20
BOTTOM_PAD_IN = 0.02

# Neutralization row height copied from 225 render (exact per-mAb vertical budget)
NEUT_ROW_H_IN = 0.85/4.0
NO_MABS_H_IN = NEUT_ROW_H_IN

def safe_normalize(mat: pd.DataFrame) -> pd.DataFrame:
    rs = mat.sum(axis=1)
    rs[rs == 0] = 1.0
    return mat.div(rs, axis=0).fillna(0)

def cell_hex(cell) -> str:
    if cell.fill and cell.fill.patternType and cell.fill.fgColor.type == "rgb":
        return "#" + cell.fill.fgColor.rgb[-6:]
    return "white"

def text_color(bg_hex: str) -> str:
    if bg_hex == "white":
        return "black"
    r = int(bg_hex[1:3], 16); g = int(bg_hex[3:5], 16); b = int(bg_hex[5:7], 16)
    return "white" if (r+g+b)/3 < 128 else "black"

def load_position_mabs(mapping_json: Path) -> dict[int, list[int]]:
    mapping_list = json.loads(mapping_json.read_text())
    pos_to_mabs = {}
    for r in mapping_list:
        p = int(r["Position"])
        s = str(r["mAbs"]).strip()
        if s in ["", "nan", "None", "none"]:
            pos_to_mabs[p] = []
        else:
            pos_to_mabs[p] = [int(x) for x in s.split(",") if str(x).strip()]
    return pos_to_mabs

def render_position(
    position: int,
    out_path: Path,
    hist_xlsx: Path,
    neut_xlsx: Path,
    aa_positions_xlsx: Path,
    mapping_json: Path,
    fmt: str = "svg",
    dpi: int = 300,
) -> Path:
    """Render one position to PNG or SVG."""

    pos_to_mabs = load_position_mabs(mapping_json)
    mabs = pos_to_mabs.get(position, [])

    # ---- histogram data ----
    df = pd.read_excel(hist_xlsx)
    sub = df[df["Position"] == position].copy()
    if sub.empty:
        raise ValueError(f"No histogram data found for position {position}")
    sub = sub[sub["SetName"] == sub["SetName"].iloc[0]].copy()

    # Coerce AA labels: any NA / N/A / blank / '-' / non-string -> "Other"
    def _coerce_aa(x):
        if pd.isna(x):
            return "Other"
        s = str(x).strip()
        if s == "-" or s.lower() in ["nan", "na", "n/a", "none", ""]:
            return "Other"
        return s

    years = np.arange(YEAR_MIN, YEAR_MAX+1)
    yrs_use = [c for c in sub.columns if isinstance(c,(int,np.integer)) and YEAR_MIN <= c <= YEAR_MAX]

    aas = sub["Amino acid"].apply(_coerce_aa).tolist()
    # Columns must be unique; duplicates can occur after coercion (e.g., many rows -> "Other")
    cols = sorted(set(aas), key=lambda z: (z == "Other", z))
    mat = pd.DataFrame(0.0, index=years, columns=cols)

    for _, rr in sub.iterrows():
        aa = _coerce_aa(rr["Amino acid"])
        # Sum if multiple rows map to the same AA label (e.g., collapsed to "Other")
        mat.loc[yrs_use, aa] += pd.to_numeric(rr[yrs_use], errors="coerce").fillna(0).values

    mat = safe_normalize(mat)

    other = [c for c in mat.columns if c.lower()=="other"]
    main  = [c for c in mat.columns if c.lower()!="other"]
    aa_order = sorted(main, key=lambda c: mat[c].mean(), reverse=True) + other

    # ---- neutralization fills + mutation letters ----
    faces = []
    mutations = {}
    asterisk_mabs = []

    if len(mabs) > 0:
        wb = openpyxl.load_workbook(neut_xlsx, data_only=False)
        ws = wb.active
        cluster_cols = {ws.cell(2,c).value:c for c in range(1,ws.max_column+1)
                        if ws.cell(2,c).value in CLUSTER_NAMES}

        def find_row(m: int):
            for rr in range(3, ws.max_row+1):
                v = ws.cell(rr, 2).value
                if v is not None:
                    try:
                        if abs(int(v)) == abs(m):
                            return rr
                    except:
                        pass
            return None

        rows = {m: find_row(m) for m in mabs}
        faces = [[cell_hex(ws.cell(rows[m], cluster_cols[c])) for c in CLUSTER_NAMES] for m in mabs]

        aa_df = pd.read_excel(aa_positions_xlsx)
        for _, rr in aa_df.iterrows():
            if int(rr["Mutation position"]) != position:
                continue
            m_raw = rr["mAB ID"]
            try:
                m_parsed = int(str(m_raw).strip())
            except Exception:
                continue
            # Treat negative/positive IDs as equivalent (e.g., -117 == 117)
            m = next((mm for mm in mabs if abs(mm) == abs(m_parsed)), None)
            if m is None:
                continue
            strain = str(rr["Virus strain"]).strip()
            mut = str(rr["Point mutation"]).strip()
            if not mut:
                continue
            wt = mut[0]; mut_aa = mut[-1]
            if strain not in CLUSTER_NAMES:
                continue
            idx = CLUSTER_NAMES.index(strain)
            if idx == len(CLUSTER_NAMES) - 1:
                asterisk_mabs.append(m)
                continue
            mutations[m] = (idx, wt, mut_aa)

    # ---- adaptive height ----
    neut_h = (NEUT_ROW_H_IN * len(mabs)) if len(mabs) > 0 else NO_MABS_H_IN
    fig_h = HIST_H_IN + CLUSTER_H_IN + neut_h + YEAR_H_IN + 3*GAP_IN + TOP_PAD_IN + BOTTOM_PAD_IN

    fig = plt.figure(figsize=(8.0, fig_h), dpi=dpi)

    left, hist_w = 0.08, 0.72
    legend_left = left + hist_w + 0.02
    legend_w = 0.98 - legend_left

    def hrel(h): return h / fig_h
    def ytop(y,h): return 1 - (y + h)/fig_h

    y = TOP_PAD_IN
    ax_hist = fig.add_axes([left, ytop(y,HIST_H_IN), hist_w, hrel(HIST_H_IN)])
    ax_leg  = fig.add_axes([legend_left, ytop(y,HIST_H_IN), legend_w, hrel(HIST_H_IN)])
    y += HIST_H_IN + GAP_IN

    ax_cluster = fig.add_axes([left, ytop(y,CLUSTER_H_IN), hist_w, hrel(CLUSTER_H_IN)])
    y += CLUSTER_H_IN + GAP_IN

    ax_neut = fig.add_axes([left, ytop(y,neut_h), hist_w, hrel(neut_h)])
    y += neut_h + GAP_IN

    ax_year = fig.add_axes([left, ytop(y,YEAR_H_IN), hist_w, hrel(YEAR_H_IN)])

    xlim = (YEAR_MIN-0.5, YEAR_MAX+0.5)
    x_left = CLUSTERS[0][1]-0.5
    x_right = CLUSTERS[-1][2]-0.5
    dx = 0.2*(xlim[1]-xlim[0])/(hist_w*8.0)

    # ---- histogram ----
    cum = np.zeros(len(years))
    for aa in aa_order:
        ax_hist.bar(
            years, mat[aa], bottom=cum, width=0.92,
            color=AA_COLOR_MAP.get(aa, AA_COLOR_MAP["Other"]),
            edgecolor="white", linewidth=0.25
        )
        cum += mat[aa].to_numpy()

    ax_hist.set_xlim(*xlim); ax_hist.set_ylim(0,1)
    ax_hist.set_title(str(position), fontweight="bold")
    ax_hist.set_ylabel("Proportion Amino Acid", fontweight="bold")
    ax_hist.set_xticks(np.arange(1970, YEAR_MAX+1, 5))
    ax_hist.set_xticklabels([])
    ax_hist.set_yticks([0,0.25,0.5,0.75,1.0])
    ax_hist.tick_params(direction="in")
    ax_hist.spines[['top','right']].set_visible(False)
    ax_hist.grid(True, linestyle=":", linewidth=0.8)

    # ---- legend/key ----
    ax_leg.axis("off")
    ax_leg.text(0.0, 1.02, "Amino Acid", transform=ax_leg.transAxes, fontweight="bold")
    yy = 0.96
    for aa in [a for a in aa_order if mat[a].sum() > 0]:
        ax_leg.add_patch(Rectangle((0.02, yy-0.045), 0.10, 0.07,
                                   transform=ax_leg.transAxes,
                                   facecolor=AA_COLOR_MAP.get(aa, AA_COLOR_MAP["Other"]),
                                   edgecolor="none"))
        ax_leg.text(0.16, yy-0.01, aa, transform=ax_leg.transAxes, va="center", fontsize=11)
        yy -= 0.10

    # ---- cluster strip ----
    ax_cluster.axis("off")
    ax_cluster.set_xlim(*xlim); ax_cluster.set_ylim(0,1)
    for n,s,e in CLUSTERS:
        s -= 0.5; e -= 0.5
        ax_cluster.add_patch(Rectangle((s,0), e-s, 1, facecolor=CLUSTER_COLORS[n], edgecolor="black"))
        ax_cluster.text((s+e)/2, 0.5, n[-2:], ha="center", va="center", fontsize=9)
    ax_cluster.text(x_right + dx, 0.5, "Antigenic cluster", ha="left", va="center", fontweight="bold")

    # ---- neutralization matrix + mutation letters ----
    ax_neut.axis("off")
    ax_neut.set_xlim(*xlim)

    if len(mabs) == 0:
        ax_neut.set_ylim(0,1)
        ax_neut.text((x_left + x_right)/2, 0.5, "No associated mAbs",
                     ha="center", va="center", fontweight="bold", fontsize=12)
    else:
        ax_neut.set_ylim(0, len(mabs))
        for i,m in enumerate(mabs):
            yrow = len(mabs) - 1 - i
            for j,(nm,s,e) in enumerate(CLUSTERS):
                s -= 0.5; e -= 0.5
                bg = faces[i][j]
                ax_neut.add_patch(Rectangle((s, yrow), e-s, 1,
                                            facecolor=bg, edgecolor="black", linewidth=0.6))
                if m in mutations and mutations[m][0] == j:
                    _, wt, mut_aa = mutations[m]
                    ax_neut.text((s+e)/2, yrow+0.5, wt,
                                 ha="center", va="center",
                                 fontsize=12, fontweight="bold", color=text_color(bg))
                    s2, e2 = CLUSTERS[j+1][1]-0.5, CLUSTERS[j+1][2]-0.5
                    bg2 = faces[i][j+1]
                    ax_neut.text((s2+e2)/2, yrow+0.5, mut_aa,
                                 ha="center", va="center",
                                 fontsize=12, fontweight="bold", color=text_color(bg2))

            label = f"mAb {m}" + ("*" if m in asterisk_mabs else "")
            ax_neut.text(x_left - dx, yrow+0.5, label,
                         ha="right", va="center", fontweight="bold")

        ax_neut.text(x_right + dx, len(mabs)/2, "Neutralization",
                     ha="left", va="center", fontweight="bold")

    # ---- year axis ----
    ax_year.set_xlim(*xlim); ax_year.set_ylim(0,1)
    ax_year.spines[['top','right','left']].set_visible(False)
    ax_year.set_yticks([])
    ax_year.set_xticks(np.arange(YEAR_MIN, YEAR_MAX+1, 5))
    ax_year.set_xticklabels([str(x) for x in np.arange(YEAR_MIN, YEAR_MAX+1, 5)],
                            rotation=90, fontweight="bold")
    ax_year.tick_params(direction="in", pad=1)
    ax_year.set_xlabel("Year", fontweight="bold", labelpad=2)

    fig.patch.set_alpha(0)
    out_path = out_path.with_suffix("." + fmt.lower())
    fig.savefig(out_path, transparent=True, bbox_inches="tight", pad_inches=0)
    plt.close(fig)
    return out_path

if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--position", type=int, required=True)
    ap.add_argument("--out", type=str, required=True)
    ap.add_argument("--fmt", type=str, default="svg", choices=["svg","png"])
    ap.add_argument("--hist_xlsx", type=str, required=True)
    ap.add_argument("--neut_xlsx", type=str, required=True)
    ap.add_argument("--aa_positions_xlsx", type=str, required=True)
    ap.add_argument("--mapping_json", type=str, required=True)
    args = ap.parse_args()

    render_position(
        position=args.position,
        out_path=Path(args.out),
        hist_xlsx=Path(args.hist_xlsx),
        neut_xlsx=Path(args.neut_xlsx),
        aa_positions_xlsx=Path(args.aa_positions_xlsx),
        mapping_json=Path(args.mapping_json),
        fmt=args.fmt
    )
